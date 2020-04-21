import shutil
from datetime import datetime
from zipfile import ZipFile

import openpyxl as pyxl

import Grading
import UI
from Config import cfg
from FileExecution import execution
from Grading.Text.text import StudentReport
from PYGUtils import *
from Viewer import Viewer

# this is the extension type assigned to files that don't have an extension
no_ext_msg = 'no-extension'


# TODO when this is finished I need to make all the static functions static (and make some static that aren't)
class Assignment:

    def __init__(self, assignment_name: str):
        # set assignment name and make directory object
        self.assignment_name = assignment_name

        # make dir dictionary with all directories relevant to this HW
        self.dir = {'home': join(cfg.base_directory, assignment_name)}
        for other_dir in ['key-output', 'key-source', 'results', 'student-source', 'test-cases',
                          'TEMP']:
            self.dir[other_dir] = join(self.dir['home'], other_dir)

        # self.criteria = None

    def can_run_key(self) -> bool:
        if len(os.listdir(self.dir['key-source'])) == 0:
            print(f'No source files found in {self.dir["key-source"]}')
            return False
        if len(os.listdir(self.dir['test-cases'])) == 0:
            print(f'No dirs found in {self.dir["test-cases"]}')
            print('Creating a default test case')
            os.mkdir(join(self.dir['test-cases'], 'default'))
        return True

    def clear_key(self) -> bool:
        print('[1] Overwrite current key files')
        print('[0] Don\'t overwrite current key files')
        choice = input_range(0, 1)
        if choice == 1:
            k = 0
            while True:
                k += 1
                if k > 25:
                    print(f'There was an issue clearing {self.dir["key output"]}')
                    return False
                try:
                    shutil.rmtree(self.dir['key-output'], ignore_errors=True)
                    os.mkdir(self.dir['key-output'])
                    break
                except FileExistsError:
                    continue
                except PermissionError:
                    print(f'Permission error clearing the directory {self.dir["key-output"]}'
                          f'Please close it if it is open')
                    return False
            return True
        else:
            return False

    def generate_key_files(self) -> None:
        # making sure there are source files and test cases
        if not self.can_run_key():
            return None

        # resolve overwriting prveious key output if there is one
        if len(os.listdir(self.dir['key-output'])) > 0:
            if not self.clear_key():
                return None

        # get a list of the finished files using the run key function
        out_file_list = execution.run_key(self.dir['home'])

        # check if there was any illegal code by looking at output files
        for file in out_file_list:
            if ' + ILLEGAL CODE + ' in execution.read_file(file):
                print(execution.read_file(file))
                print('\nGrading criteria cannot be created. Returning to menu...')
                return None

        # find the parts of the assignment from the outfile list
        problem_parts = Grading.Text.text.find_parts(out_file_list)
        if not problem_parts:
            problem_parts = ['default']

        generate_blank_rubric(problem_parts, join(self.dir['home'], 'rubric.ini'), self.assignment_name)
        print(f'Please fill out {join(self.dir["home"], "rubric.ini")}')
        print('in order to be able to grade the file.')

    def export_student_tester(self):
        print('export student tester')

    def choose_student_source_dir(self) -> str:
        # TODO maybe I could support having zip files here too for holding student code
        source_dirs = []
        for file in os.listdir(self.dir['student-source']):
            if os.path.isdir(join(self.dir['student-source'], file)):
                source_dirs.append(file)

        # choose which directory of student files they should use
        if len(source_dirs) == 0:
            return ''

        print('Choose a batch of student files to grade from:')
        for i, source_dir in enumerate(source_dirs):
            print(f'[{i + 1}] - {source_dir}')
        return join(self.dir['student-source'], source_dirs[input_range(1, len(source_dirs)) - 1])

    def get_ids_from_files(self, source_dir: str) -> typing.Tuple[typing.List[str], typing.Set[str]]:
        student_ids = set()
        types_found = set()
        for file in os.listdir(source_dir):
            if '.' not in file:
                types_found.add(no_ext_msg)
            else:
                student_ids.add('_'.join(file.split('_')[:3]))
                types_found.add(file.split('.')[-1])
        return sorted(student_ids), types_found

    def make_temp_dir(self, student_ids: typing.List[str]) -> None:
        # populate a directory that will get filled up and eventually zipped as the report
        if os.path.exists(self.dir['TEMP']):
            shutil.rmtree(self.dir['TEMP'])
        os.mkdir(self.dir['TEMP'])

        # make a directory for each student in the temp dir
        for student_id in student_ids:
            os.mkdir(join(self.dir['TEMP'], student_id))

    def get_types_to_move(self, types_found: typing.Set[str]) -> typing.List[str]:
        types_to_move = []
        print(f'\nSubmitted file types: {types_found}')
        print('Enter 2 to add all')
        for file_type in types_found:
            choice = input_range(0, 2,
                                 message=f'Move student .{file_type} files\n[1] - yes\n[2] - no\n[0] - yes to all\n')
            if choice == 1:
                types_to_move.append(file_type)
            elif choice == 0:
                types_to_move = list(types_found)
                break
        return types_to_move

    def move_student_files(self, source_dir: str, types_to_move: typing.List[str]) -> None:
        # I HAVE DECIDED TO REMOVE THE IDS FROM THE FILES HERE
        for file in os.listdir(source_dir):
            try:
                if file.split('.')[-1] in types_to_move:
                    file_id = '_'.join(file.split('_')[:3])
                    file_name = '_'.join(file.split('_')[3:])
                    shutil.copyfile(join(source_dir, file), join(self.dir['TEMP'], file_id, file_name))
            except IndexError:
                if no_ext_msg in types_to_move:
                    file_id = '_'.join(file.split('_')[:3])
                    file_name = '_'.join(file.split('_')[3:])
                    shutil.copyfile(join(source_dir, file), join(self.dir['TEMP'], file_id, file_name))

    def zip_report(self) -> None:
        # Note this will only go one folder deep into the temp folder, but this shouldn't be an issue
        timestamp = datetime.now().strftime('%d-%b-%Y %I-%M-%S %p')
        with ZipFile(f'{join(self.dir["results"], timestamp)}.zip', 'w') as zip_obj:
            for folder in os.listdir(self.dir['TEMP']):
                if os.path.isdir(join(self.dir['TEMP'], folder)):
                    for file in os.listdir(join(self.dir['TEMP'], folder)):
                        zip_obj.write(join(self.dir['TEMP'], folder, file), arcname=join(folder, file))
                else:
                    zip_obj.write(join(self.dir['TEMP'], folder), arcname=folder)
        shutil.rmtree(self.dir['TEMP'])

    def grade_student_code(self) -> None:
        # TODO need to load the ini file (make a function in tools) and catch any errors loading it

        source_dir = self.choose_student_source_dir()

        if source_dir == '':
            print('There are no student source directories in')
            print(join(cfg.base_directory, self.dir['student-source']))
            return None

        # get the student ids from the files in the selected student source directory
        student_ids, types_found = self.get_ids_from_files(source_dir)

        self.make_temp_dir(student_ids)

        # make a list of file types to move
        types_to_move = self.get_types_to_move(types_found)

        # move the student source files into the temp folder
        self.move_student_files(source_dir, types_to_move)

        # run the student code on a bunch of threads and leaves the output.txt s
        out_files = execution.run_students(self.dir['home'])

        grades = Grading.Text.text.grade_students(self.dir['home'], out_files)

        self.generate_grade_report(grades)

        # make a zip file in results and copy everything from temp into it
        self.zip_report()

    def generate_grade_report(self, grades: typing.List[StudentReport]) -> None:
        wb = pyxl.Workbook()
        grade_sheet = wb.create_sheet('Grades')
        grade_sheet.cell(1, 1, 'Last Name')
        grade_sheet.cell(1, 2, 'First Name')
        grade_sheet.cell(1, 3, 'NetID')
        grade_sheet.cell(1, 4, 'Total - Weighted')
        grade_sheet.cell(1, 5, 'Total')
        index = 6
        reader = configparser.ConfigParser()
        reader.read(join(self.dir['home'], 'rubric.ini'))
        weighted_total = reader.getint('Assignment', 'total_weight')
        parts = get_assignment_parts(self.dir['home'])
        part_weights = {}
        possible_points = 0
        for part in reader.options('Parts'):
            part_weights[part] = reader.getint('Parts', part)
            possible_points += part_weights[part]
        test_cases = get_assignment_test_cases(self.dir['home'])
        for part in parts:
            for test_case in test_cases:
                grade_sheet.cell(1, index, f'{part} - {test_case}')
                index += 1

        for i, student in enumerate(grades):
            grade_sheet.cell(i + 2, 1, student.identity['last'])
            grade_sheet.cell(i + 2, 2, student.identity['first'])
            grade_sheet.cell(i + 2, 3, student.identity['id'])
            index = 6
            student_total = 0
            for part in parts:
                for test_case in test_cases:
                    grade_sheet.cell(i + 2, index, student.test_cases[test_case][part])
                    index += 1
                    if student.test_cases[test_case][part]:
                        student += part_weights[part.replace(' ', '_')]
            grade_sheet.cell(i + 2, 5, student_total)
            grade_sheet.cell(i + 2, 4, round(student_total / possible_points * weighted_total, cfg.score_decimals))

        stats_sheet = wb.create_sheet('Statistics')

        wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
        wb.save(join(self.dir['TEMP'], 'grade_report.xlsx'))

    def view_grading_report(self) -> None:
        # TODO make reporter object and pass it to the UI
        # TODO this should also pop up a seperate window with the xlsx report (this would have to be from the program)

        if len(os.listdir(self.dir['results'])) == 0:
            print(f'No reports have been generated for {self.assignment_name} yet')
            return None

        zips = []
        for file in os.listdir(self.dir['results']):
            if file.endswith('.zip'):
                zips.append(file)
        print('Choose report to view:')
        for i, file in enumerate(zips):
            print(f'[{i}] {file}')
        print('[0] Cancel')
        choice = input_range(0, len(zips))
        if choice == 0:
            return None

        viewer = Viewer(self.dir['results'], zips[choice - 1])
        UI.start_ui(viewer)
