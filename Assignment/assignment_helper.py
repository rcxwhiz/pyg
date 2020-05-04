import configparser
import os
import shutil
import typing
from datetime import datetime
from os.path import join
from zipfile import ZipFile

import openpyxl as pyxl

import PYGUtils as putil
from Config import cfg
from Grading.Text.text import StudentReport

no_ext_msg = 'no-extension'


def can_run_key(assignment_dir: str) -> bool:
    if not os.path.exists(join(assignment_dir, 'key-source')):
        print('No source directory found, creating empty one...')
        os.mkdir(join(assignment_dir, 'key-source'))
    if len(os.listdir(join(assignment_dir, 'key-source'))) == 0:
        print(f'No source files found in {join(assignment_dir, "key-source")}')
        return False
    if len(os.listdir(join(assignment_dir, 'test-cases'))) == 0:
        print('No test cases found, creating default test case...')
        os.mkdir(join(assignment_dir, 'test-cases', 'default'))
    return True


def clear_key(assignment_dir: str) -> bool:
    if len(os.listdir(join(assignment_dir, 'key-output'))) > 0:
        print('[1] Overwrite current key output files')
        print('[0] Cancel')
        choice = putil.input_range(0, 1)
        if choice == 0:
            return False
        k = 0
        while True:
            k += 1
            if k > 25:
                print(f'There was an issue clearing {join(assignment_dir, "key-output")}')
                return False
            try:
                shutil.rmtree(join(assignment_dir, 'key-output'), ignore_errors=True)
                os.mkdir(join(assignment_dir, 'key-output'))
                break
            except FileExistsError:
                continue
            except PermissionError:
                print(f'Permission denied to remove directory {join(assignment_dir, "key-output")}')
                print('Please close it if it is open')
                return False
    return True


def get_ids_from_files(source_dir: str) -> typing.Tuple[typing.List[str], typing.Set[str]]:
    student_ids = set()
    types_found = set()

    for file in os.listdir(source_dir):
        if '.' not in file:
            types_found.add(no_ext_msg)
        elif putil.get_id(file) is not None:
            student_ids.add('_'.join(putil.get_id(file)))
            types_found.add(file.split('.')[-1])
    return sorted(student_ids), types_found


def makedir(assignment_dir: str, student_ids: typing.List[str]) -> None:
    if os.path.exists(join(assignment_dir, 'TEMP')):
        shutil.rmtree(join(assignment_dir, 'TEMP'))
    os.mkdir(join(assignment_dir, 'TEMP'))

    for student_id in student_ids:
        os.mkdir(join(assignment_dir, 'TEMP', student_id))


def get_types_to_move(types_found: typing.Set[str]) -> typing.List[str]:
    if len(types_found) == 1:
        return list(types_found)
    types_to_move = []
    print(f'\nSubmitted file types: {types_found}')
    for file_type in types_found:
        choice = putil.input_range(0, 2,
                                   message=f'Move student .{file_type} files\n[1] - yes\n[2] - no\n[0] - yes to all\n')
        if choice == 1:
            types_to_move.append(file_type)
        elif choice == 0:
            types_to_move = list(types_found)
            break
    return types_to_move


def move_student_files(assignment_dir: str, source_dir: str, types_to_move: typing.List[str]) -> None:
    for file in os.listdir(source_dir):
        try:
            if ('.' not in file and no_ext_msg in types_to_move) or file.split('.')[-1] in types_to_move:
                file_id = '_'.join(file.split('_')[:3])
                file_name = '_'.join(file.split('_'[3:]))
                shutil.copyfile(join(source_dir, file), join(assignment_dir, 'TEMP', file_id, file_name))
        except IndexError:
            if '.' in file:
                print(f'There was an issue getting the id and name from {file}')


def generate_grade_report(assignment_dir: str, grades: typing.List[StudentReport]) -> None:
    wb = pyxl.Workbook()
    grade_sheet = wb.create_sheet('Grades')
    grade_sheet.cell(1, 1, 'Last Name')
    grade_sheet.cell(1, 2, 'First Name')
    grade_sheet.cell(1, 3, 'NetID')
    grade_sheet.cell(1, 4, 'Total - Weighted')
    grade_sheet.cell(1, 5, 'Total')
    index = 6
    reader = configparser.ConfigParser()
    reader.read(join(assignment_dir, 'rubric.ini'))
    weighted_total = reader.getint('Assignment', 'total_weight')
    parts = putil.get_assignment_parts(assignment_dir)
    part_weights = {}
    possible_points = 0
    for part in reader.options('Parts'):
        part_weights[part] = reader.getint('Parts', part)
        possible_points += part_weights[part]
    test_cases = putil.get_assignment_test_cases(assignment_dir)
    for part in parts:
        for test_case in test_cases:
            grade_sheet.cell(1, index, f'{part} - {test_case}')
            index += 1

    for i, student in enumerate(grades):
        grade_sheet.cell(i + 2, 1, student.identity['last'])
        grade_sheet.cell(i + 2, 2, student.identity['first'])
        grade_sheet.cell(i + 2, 3, student.identity['id'])
        index = 6
        student_total = 0
        for part in parts:
            for test_case in test_cases:
                grade_sheet.cell(i + 2, index, student.test_cases[test_case][part])
                index += 1
                if student.test_cases[test_case][part]:
                    student += part_weights[part.replace(' ', '_')]
        grade_sheet.cell(i + 2, 5, student_total)
        grade_sheet.cell(i + 2, 4, round(student_total / possible_points * weighted_total, cfg.score_decimals))

    # TODO make statistics sheet
    # stats_sheet = wb.create_sheet('Statistics')

    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
    wb.save(join(assignment_dir, 'grade_report.xlsx'))


def zip_report(assignment_dir: str) -> None:
    temp_dir = join(assignment_dir, 'TEMP')
    timestamp = datetime.now().strftime('%d-%b-%Y %I-%M-%S %p')
    with ZipFile(f'{join(assignment_dir, "results", timestamp)}.zip', 'w') as zip_obj:
        for folder in os.listdir(temp_dir):
            if os.path.isdir(join(temp_dir, folder)):
                for file in os.listdir(join(temp_dir, folder)):
                    zip_obj.write(join(temp_dir, folder, file), arcname=join(folder, file))
            else:
                zip_obj.write(join(temp_dir, folder), arcname=folder)
    shutil.rmtree(temp_dir)
